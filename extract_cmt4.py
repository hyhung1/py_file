from apify_client import ApifyClient
from datetime import datetime
import openpyxl
import json
import os
import sys
import requests
import hashlib
from pathlib import Path
import glob
from operator import itemgetter

# Configuration settings
CONFIG = {
    "API_KEY": "apify_api_aiwzjla0jREK3paH0j7cmhrZlgjPWB0mXFRe",
    "MAX_COMMENTS": 20,         # Maximum comments to extract from TikTok
    "TOP_COMMENTS": 6,          # Top comments to keep after sorting by engagement
    "DEFAULT_JSON_FILE": "./QUANAN_alpha/q10/quán_ngon_quận_10_upd.json",
    "OUTPUT_BASE_FOLDER": "./processed_data",
    "COMMENT_ACTOR_ID": "XomSRf7d0qf3mVj1y",  # TikTok comment extraction actor ID
    "REPLY_WEIGHT": 2.0,        # Weight for reply count in engagement score (higher priority)
    "LIKE_WEIGHT": 1.0,         # Weight for like count in engagement score
}

def download_avatar(avatar_url, save_dir, username):
    """Download avatar image from URL"""
    try:
        # Basic URL validation
        if not avatar_url or not avatar_url.startswith(('http://', 'https://')):
            print(f"  Invalid avatar URL for {username}")
            return None
            
        url_hash = hashlib.md5(avatar_url.encode()).hexdigest()[:8]
        safe_username = ''.join(c for c in username if c.isalnum())[:20]
        
        # Initially create path without extension
        base_path = Path(save_dir) / f"{safe_username}_{url_hash}"
        
        # Check if any version of the file exists
        existing_files = list(Path(save_dir).glob(f"{safe_username}_{url_hash}.*"))
        if existing_files:
            print(f"  Avatar already exists for {username}")
            return str(existing_files[0])
            
        # Download with timeout
        response = requests.get(avatar_url, stream=True, timeout=10)
        response.raise_for_status()  # Raise exception for bad status codes
        
        # Get content type and determine extension
        content_type = response.headers.get('content-type', '').lower()
        if 'image' not in content_type:
            print(f"  Invalid content type for {username}: {content_type}")
            return None
            
        # Map content type to extension
        extension_map = {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'image/webp': '.webp',
            'image/gif': '.gif'
        }
        extension = extension_map.get(content_type, '.jpg')
        save_path = base_path.with_suffix(extension)
        
        # Ensure directory exists
        save_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Download and save the file
        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    
        print(f"  Downloaded avatar for {username}")
        return str(save_path)
        
    except requests.exceptions.Timeout:
        print(f"  Timeout downloading avatar for {username}")
        return None
    except requests.exceptions.RequestException as e:
        print(f"  Network error downloading avatar for {username}: {str(e)}")
        return None
    except Exception as e:
        print(f"  Error downloading avatar for {username}: {str(e)}")
        return None

def calculate_engagement_score(comment):
    """
    Calculate an engagement score for a comment based on replyCount and likeCount.
    Prioritizes replyCount by applying a higher weight.
    """
    reply_count = comment.get("replyCount", 0) or 0
    like_count = comment.get("likeCount", 0) or 0
    
    # Calculate weighted score
    score = (reply_count * CONFIG["REPLY_WEIGHT"]) + (like_count * CONFIG["LIKE_WEIGHT"])
    return score

def extract_tiktok_comments(api_key, url, max_items=80, top_comments=5, output_file=None, avatar_dir=None):
    """Extract comments from a TikTok video and filter top comments by engagement"""
    client = ApifyClient(api_key)

    run_input = {
        "startUrls": [url],
        "includeReplies": True,
        "maxItems": max_items,
        "customMapFunction": "(object) => { return {...object} }",
    }

    print(f"Extracting up to {max_items} comments from {url}...")
    run = client.actor(CONFIG["COMMENT_ACTOR_ID"]).call(run_input=run_input)

    all_comments = []
    print("Processing comments...")
    
    if avatar_dir:
        Path(avatar_dir).mkdir(parents=True, exist_ok=True)

    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        created_at = item.get("createdAt", "")
        formatted_date = ""
        if created_at:
            date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            formatted_date = date_obj.strftime('%d-%m-%Y')
        
        filtered_item = {
            "text": item.get("text"),
            "createdAt": formatted_date,
            "likeCount": item.get("likeCount", 0),
            "replyCount": item.get("replyCount", 0),
            "isAuthorLiked": item.get("isAuthorLiked")
        }
        
        if "user" in item and item["user"]:
            user_data = item["user"]
            filtered_item["username"] = user_data.get("username")
            filtered_item["displayName"] = user_data.get("displayName")
            filtered_item["bio"] = user_data.get("bio")
            filtered_item["avatarUrl"] = user_data.get("avatarUrl")
            
            if avatar_dir and filtered_item["avatarUrl"] and filtered_item["username"]:
                # Add retry logic for failed downloads
                max_retries = 3
                avatar_path = None
                for retry in range(max_retries):
                    avatar_path = download_avatar(
                        filtered_item["avatarUrl"],
                        avatar_dir,
                        filtered_item["username"]
                    )
                    if avatar_path:
                        break
                    print(f"  Retry {retry + 1}/{max_retries} for {filtered_item['username']}")
                filtered_item["avatar_local_path"] = avatar_path

        # Calculate engagement score
        filtered_item["engagement_score"] = calculate_engagement_score(filtered_item)
        all_comments.append(filtered_item)

    # Sort comments by engagement score (descending)
    all_comments.sort(key=itemgetter("engagement_score"), reverse=True)
    
    # Get top comments based on engagement score
    top_comments_data = all_comments[:top_comments] if top_comments < len(all_comments) else all_comments
    
    print(f"Found {len(all_comments)} comments, keeping top {len(top_comments_data)} by engagement score")
    
    # Create a folder for all comments too
    if output_file:
        all_comments_file = Path(output_file).with_name(f"{Path(output_file).stem}_all.xlsx")
        save_comments_to_excel(all_comments, str(all_comments_file))
        print(f"All {len(all_comments)} comments saved to {all_comments_file}")
    
    # Save top comments to the main Excel file
    if top_comments_data and output_file:
        save_comments_to_excel(top_comments_data, output_file)
    
    return len(top_comments_data), top_comments_data, len(all_comments)

def save_comments_to_excel(comments_data, output_file):
    """Save comments data to an Excel file"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TikTok Comments"

    headers = ["Text", "Created At", "Like Count", "Reply Count", "Is Author Liked", 
              "Username", "Display Name", "Bio", "Avatar URL", "Avatar Local Path", "Engagement Score"]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    for row, data in enumerate(comments_data, 2):
        sheet.cell(row=row, column=1).value = data.get("text", "")
        sheet.cell(row=row, column=2).value = data.get("createdAt", "")
        sheet.cell(row=row, column=3).value = data.get("likeCount", 0)
        sheet.cell(row=row, column=4).value = data.get("replyCount", 0)
        sheet.cell(row=row, column=5).value = "Yes" if data.get("isAuthorLiked") else "No"
        sheet.cell(row=row, column=6).value = data.get("username", "")
        sheet.cell(row=row, column=7).value = data.get("displayName", "")
        sheet.cell(row=row, column=8).value = data.get("bio", "")
        sheet.cell(row=row, column=9).value = data.get("avatarUrl", "")
        sheet.cell(row=row, column=10).value = data.get("avatar_local_path", "")
        sheet.cell(row=row, column=11).value = data.get("engagement_score", 0)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Use pathlib for output file
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        workbook.save(output_path)
        print(f"Data saved to {output_path}")
    except Exception as e:
        print(f"ERROR SAVING EXCEL FILE: {str(e)}")
        raise

def get_parent_folder_name(folder_path):
    """Get the name of the parent folder"""
    path = Path(folder_path)
    return path.name

def process_json_file(json_file, api_key, max_comments=80, top_comments=5, output_base_folder=None):
    """Process all restaurants in the JSON file"""
    try:
        json_path = Path(json_file)
        if not json_path.exists():
            print(f"ERROR: JSON file not found: {json_path}")
            return
            
        with open(json_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        print(f"Found {len(data)} restaurants in the JSON file")
        print(f"Will extract up to {max_comments} comments per restaurant and keep top {top_comments} by engagement")
        
        for index, item in enumerate(data):
            print(f"\nProcessing restaurant {index+1}/{len(data)}")
            
            try:
                restaurant_name = item.get('eat_name', 'Unknown')
                usn_time = item.get('usn_time', '')
                
                print(f"Restaurant: {restaurant_name}")
                print(f"USN_TIME: {usn_time}")
                
                # Determine comment paths based on output_base_folder or from JSON
                if output_base_folder:
                    # If we have an output base folder, construct paths using usn_time
                    # Create a sanitized folder name
                    base_folder = Path(output_base_folder)
                    
                    # Handle both cases - with or without usn_time
                    if usn_time:
                        safe_folder_name = ''.join(c for c in usn_time if c.isalnum() or c in ('_', '-'))[:50]
                    else:
                        safe_folder_name = ''.join(c for c in restaurant_name if c.isalnum() or c in ('_', '-'))[:50]
                    
                    restaurant_folder = base_folder / safe_folder_name
                    comments_path = restaurant_folder / "comments"
                    user_cover_img = restaurant_folder / "comments" / "user_cover_img"
                    
                    # Use the parent folder name (usn_time or restaurant name folder)
                    excel_filename = f"{restaurant_folder.name}.xlsx"
                    
                    print(f"  Using output folder structure at: {restaurant_folder}")
                else:
                    # Use paths from JSON
                    comments_path = Path(item.get('comments_path', ''))
                    user_cover_img = Path(item.get('user_cover_img', ''))
                    
                    if not comments_path.is_absolute():
                        comments_path = Path.cwd() / comments_path
                        
                    if not user_cover_img.is_absolute():
                        user_cover_img = Path.cwd() / user_cover_img
                    
                    # Get the parent folder name (the folder above comments)
                    parent_folder = comments_path.parent
                    excel_filename = f"{parent_folder.name}.xlsx"
                    
                    print(f"  Using JSON-defined paths")
                
                print(f"  Comments path: {comments_path}")
                print(f"  Avatar path: {user_cover_img}")
                
                # Create directories if they don't exist
                comments_path.mkdir(parents=True, exist_ok=True)
                user_cover_img.mkdir(parents=True, exist_ok=True)
                
                post_page = item.get('postPage')
                if not post_page:
                    print(f"  ERROR: TikTok URL (postPage) not found for {restaurant_name}")
                    continue
                
                print(f"  TikTok URL: {post_page}")
                
                # Create Excel path with the parent folder name
                excel_path = comments_path / excel_filename
                
                print(f"  Will save top comments to: {excel_path}")
                
                # Extract comments and save to Excel, also download avatars
                top_count, _, total_count = extract_tiktok_comments(
                    api_key=api_key,
                    url=post_page,
                    max_items=max_comments,
                    top_comments=top_comments,
                    output_file=str(excel_path),
                    avatar_dir=str(user_cover_img)
                )
                
                print(f"  SUCCESS: Extracted {total_count} comments, saved top {top_count} for {restaurant_name}")
            
            except Exception as e:
                print(f"  ERROR processing restaurant {restaurant_name}: {str(e)}")
                continue  # Added continue to process next restaurant even if one fails
        
        print("\nProcessing complete!")
    
    except Exception as e:
        print(f"ERROR: Failed to process JSON file: {str(e)}")

def process_folder_structure(output_base_folder):
    """Find all JSON files in the folder structure and process them"""
    try:
        # Ensure the folder exists
        base_folder = Path(output_base_folder)
        if not base_folder.exists():
            print(f"ERROR: Base folder not found: {base_folder}")
            return
            
        # Search recursively for JSON files that might have the right structure
        json_files = glob.glob(str(base_folder / "**" / "*_processed.json"), recursive=True)
        json_files.extend(glob.glob(str(base_folder / "**" / "*_addurl.json"), recursive=True))
        
        if not json_files:
            print(f"No suitable JSON files found in {base_folder}")
            return
            
        print(f"Found {len(json_files)} JSON files to process")
        
        for json_file in json_files:
            print(f"\n{'='*60}")
            print(f"Processing file: {json_file}")
            print(f"{'='*60}")
            
            # Get the parent folder of this JSON file to use as the output base
            json_parent = Path(json_file).parent
            
            process_json_file(
                json_file=json_file,
                api_key=CONFIG["API_KEY"],
                max_comments=CONFIG["MAX_COMMENTS"],
                top_comments=CONFIG["TOP_COMMENTS"],
                output_base_folder=json_parent  # Use the JSON's parent folder
            )
    
    except Exception as e:
        print(f"ERROR processing folder structure: {str(e)}")

def main():
    # Check if we have command line arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == '--help' or sys.argv[1] == '-h':
            print("Usage: python script.py [json_file] [output_folder] [max_comments] [top_comments]")
            print("  json_file:     Path to a specific JSON file to process")
            print("  output_folder: Base folder containing JSON files and subfolders")
            print("  max_comments:  Maximum comments to extract per video (default: 20)")
            print("  top_comments:  Top comments to keep after sorting by engagement (default: 5)")
            sys.exit(0)
        
        # First argument is the JSON file
        CONFIG["DEFAULT_JSON_FILE"] = sys.argv[1]
        
        # Second argument (if provided) is the output base folder
        if len(sys.argv) > 2:
            CONFIG["OUTPUT_BASE_FOLDER"] = sys.argv[2]
            
        # Third argument (if provided) is max comments
        if len(sys.argv) > 3:
            try:
                CONFIG["MAX_COMMENTS"] = int(sys.argv[3])
            except ValueError:
                print(f"Warning: Invalid value for max_comments: {sys.argv[3]}. Using default: {CONFIG['MAX_COMMENTS']}")
                
        # Fourth argument (if provided) is top comments
        if len(sys.argv) > 4:
            try:
                CONFIG["TOP_COMMENTS"] = int(sys.argv[4])
            except ValueError:
                print(f"Warning: Invalid value for top_comments: {sys.argv[4]}. Using default: {CONFIG['TOP_COMMENTS']}")
    
    print(f"Current working directory: {Path.cwd()}")
    print(f"Max comments to extract: {CONFIG['MAX_COMMENTS']}")
    print(f"Top comments to keep: {CONFIG['TOP_COMMENTS']}")
    
    # If output folder is provided, process all JSON files in the structure
    if CONFIG["OUTPUT_BASE_FOLDER"]:
        print(f"Processing folder structure: {CONFIG['OUTPUT_BASE_FOLDER']}")
        process_folder_structure(CONFIG["OUTPUT_BASE_FOLDER"])
    else:
        # Otherwise, process the single JSON file
        json_path = Path(CONFIG["DEFAULT_JSON_FILE"])
        
        if not json_path.exists():
            print(f"ERROR: JSON file not found: {json_path}")
            sys.exit(1)
        else:
            print(f"JSON file found: {json_path.absolute()}")
        
        process_json_file(
            json_file=str(json_path),
            api_key=CONFIG["API_KEY"],
            max_comments=CONFIG["MAX_COMMENTS"],
            top_comments=CONFIG["TOP_COMMENTS"]
        )

if __name__ == "__main__":
    main()