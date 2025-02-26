from apify_client import ApifyClient
from typing import Dict, List, Any
import json
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Centralized configuration dictionary
CONFIG = {
    # API credentials
    "API_KEY": "apify_api_aiwzjla0jREK3paH0j7cmhrZlgjPWB0mXFRe",
    
    # Search parameters
    "SEARCH_KEYWORDS": ["quán ngon quận 11"],
    "MAX_ITEMS": 150,
    "DATE_RANGE": "DEFAULT",
    "LOCATION": "VN",
    
    # Project structure
    "PROJECT_FOLDER": "QUANAN_new_150",
    "PROJECT_DIST": "q11",
    
    # Excel configuration
    "EXCEL_HEADERS": ["usn_time", "postPage", "eat_name", "eat_addr", "open_time", "menu"],
    "HEADER_COLOR": "DDEBF7",
    
    # Actor configuration
    "TIKTOK_SEARCH_ACTOR_ID": "5K30i8aFccKNF5ICs"
}

def search_tiktok_videos() -> List[Dict[str, Any]]:
    """
    Search TikTok videos based on keywords using Apify API and extract specific fields
    
    Returns:
        List[Dict]: List of dictionaries containing extracted video information
    """
    # Initialize the ApifyClient
    client = ApifyClient(CONFIG["API_KEY"])
    
    # Format the search term for the filename
    search_term = " ".join(CONFIG["SEARCH_KEYWORDS"])
    search_slug = search_term.replace(" ", "_")
    
    # Setup directories
    json_data_dir = f"{CONFIG['PROJECT_FOLDER']}/{CONFIG['PROJECT_DIST']}"
    os.makedirs(json_data_dir, exist_ok=True)
    output_file = f"{json_data_dir}/{search_slug}.json"
    
    # Prepare the Actor input
    run_input = {
        "maxItems": CONFIG["MAX_ITEMS"],
        "keywords": CONFIG["SEARCH_KEYWORDS"],
        "dateRange": CONFIG["DATE_RANGE"],
        "location": CONFIG["LOCATION"],
        "customMapFunction": "(object) => { return {...object} }"
    }
    
    print(f"Searching TikTok for: {search_term}")
    print(f"Maximum items: {CONFIG['MAX_ITEMS']}")
    
    # Run the Actor and wait for it to finish
    run = client.actor(CONFIG["TIKTOK_SEARCH_ACTOR_ID"]).call(run_input=run_input)
    
    results = []
    
    # Fetch and process Actor results
    print("Processing search results...")
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        # Extract only the requested fields
        extracted_data = {
            "title": item.get("title"),
            "views": item.get("views"),
            "likes": item.get("likes"),
            "comments": item.get("comments"),
            "shares": item.get("shares"),
            "bookmarks": item.get("bookmarks"),
            "hashtags": item.get("hashtags"),
            "uploadedAt": item.get("uploadedAt"),
            "uploadedAtFormatted": item.get("uploadedAtFormatted"),
            "channel": {
                "name": item.get("channel", {}).get("name"),
                "username": item.get("channel", {}).get("username")
            },
            "postPage": item.get("postPage")
        }
        
        # Format date and create identifier (similar to author_time in original code)
        upload_time = item.get("uploadedAt")
        if upload_time:
            dt = datetime.fromtimestamp(upload_time)
            formatted_date = dt.strftime('%Y_%d_%m')
            channel_name = extracted_data["channel"]["username"]
            extracted_data["usn_time"] = f"{channel_name}_{formatted_date}"
        
        results.append(extracted_data)
    
    # Save to JSON file
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=4)
    
    print(f"Data has been saved to {output_file}")
    print(f"Retrieved {len(results)} results")
    
    # Create Excel file
    excel_file = f"{json_data_dir}/{search_slug}.xlsx"
    create_excel_file(results, excel_file)
    print(f"Excel file has been created: {excel_file}")
    
    return results, output_file, excel_file

def create_excel_file(data: List[Dict], filename: str):
    """
    Create an Excel file with specific fields from the TikTok data
    
    Args:
        data (List[Dict]): List of dictionaries containing video data
        filename (str): Path to save the Excel file
    """
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TikTok Data"
    
    # Get headers from config
    headers = CONFIG["EXCEL_HEADERS"]
    
    # Style for header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color=CONFIG["HEADER_COLOR"], 
                             end_color=CONFIG["HEADER_COLOR"], 
                             fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, video in enumerate(data, 2):
        # Fill in fields from the JSON data
        ws.cell(row=row_num, column=1, value=video.get("usn_time", ""))
        ws.cell(row=row_num, column=2, value=video.get("postPage", ""))
        
        # Leave other fields empty for manual filling
        # eat_name (column 3), eat_addr (column 4), open_time (column 5), menu (column 6)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)

def save_to_json(data: List[Dict], filename: str = None) -> str:
    """
    Save the search results to a JSON file
    
    Args:
        data (List[Dict]): Data to save
        filename (str, optional): Filename to use. If None, generates a timestamp-based filename
        
    Returns:
        str: Path to the saved JSON file
    """
    if filename is None:
        # Generate timestamp-based filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tiktok_search_results_{timestamp}.json"
    
    # Ensure filename ends with .json
    if not filename.endswith('.json'):
        filename += '.json'
    
    # Save data to JSON file
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    return filename

def print_search_results(videos):
    """Print formatted search results to console"""
    for i, video in enumerate(videos, 1):
        print(f"\nVideo {i}:")
        print(f"Title: {video['title']}")
        print(f"Channel: {video['channel']['name']} (@{video['channel']['username']})")
        print(f"Stats: {video['views']} views, {video['likes']} likes, {video['comments']} comments")
        print(f"Engagement: {video['shares']} shares, {video['bookmarks']} bookmarks")
        print(f"Uploaded: {video.get('uploadedAtFormatted', 'N/A')}")
        print(f"Hashtags: {', '.join(video['hashtags']) if isinstance(video['hashtags'], list) else video['hashtags']}")
        print(f"URL: {video['postPage']}")
        print(f"Identifier: {video.get('usn_time', 'N/A')}")
        print("-" * 50)

def main():
    """Main function to execute the TikTok search"""
    print(f"Starting TikTok search for: {' '.join(CONFIG['SEARCH_KEYWORDS'])}")
    print(f"Project folder: {CONFIG['PROJECT_FOLDER']}/{CONFIG['PROJECT_DIST']}")
    print(f"Max items: {CONFIG['MAX_ITEMS']}")
    
    # Search for TikTok videos
    videos, json_file, excel_file = search_tiktok_videos()
    
    # Print the results
    print_search_results(videos)
    
    print(f"\nResults have been saved to:")
    print(f"- JSON: {json_file}")
    print(f"- Excel: {excel_file}")

if __name__ == "__main__":
    main()