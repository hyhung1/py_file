import json
import os
import cv2
import requests
import re
import glob
from pprint import pprint
from apify_client import ApifyClient

# Configuration - put all variables in one place
CONFIG = {
    "API_KEY": "apify_api_85rFJrdop3zajntC7oFQy2DabXYYjH3hJMB5",
    "INPUT_FOLDER": "./cac_quanan_q10/json_xlsx",  # Folder containing JSON files
    "OUTPUT_BASE_FOLDER": "./processed_data",  # Base output folder
    "FRAME_INTERVAL": 3,  # Extract frames every 3 seconds
    "SUBFOLDERS": ["vid", "img", "final_imgs", "cover_img", "comments", 
                   "comments/filter_cmt", "comments/user_cover_img"]
}

def extract_frames(video_path, output_folder, interval=3):
    """Extract frames from a video at specified interval (seconds)"""
    # Create a temporary folder for extraction
    import tempfile
    temp_dir = tempfile.mkdtemp()
    print(f"Created temporary directory for frame extraction: {temp_dir}")
    
    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Open the video file
    video_capture = cv2.VideoCapture(video_path)

    # Get the frames per second (fps) of the video
    fps = int(video_capture.get(cv2.CAP_PROP_FPS))
    total_frames = int(video_capture.get(cv2.CAP_PROP_FRAME_COUNT))
    duration = total_frames // fps
    
    # Track the frame paths
    frame_paths = []

    # Loop through the video at specified intervals
    for i in range(0, duration, interval):
        # Set the video position to the i-th second
        video_capture.set(cv2.CAP_PROP_POS_MSEC, i * 1000)
        success, frame = video_capture.read()
        if success:
            # Save the frame as an image file in temp directory
            frame_filename = f"frame_{i}.jpg"
            temp_frame_path = os.path.join(temp_dir, frame_filename)
            cv2.imwrite(temp_frame_path, frame)
            
            # Copy the frame to the output folder
            final_frame_path = os.path.join(output_folder, frame_filename)
            import shutil
            shutil.copy2(temp_frame_path, final_frame_path)
            
            frame_paths.append(final_frame_path)
        else:
            break

    # Release the video capture object
    video_capture.release()
    
    # Clean up the temporary directory
    import shutil
    shutil.rmtree(temp_dir)
    print(f"Removed temporary directory: {temp_dir}")
    
    return frame_paths

def download_mp4(url, output_path):
    """Download MP4 file from URL"""
    response = requests.get(url, stream=True)
    if response.status_code == 200:
        with open(output_path, 'wb') as file:
            for chunk in response.iter_content(chunk_size=1024):
                if chunk:
                    file.write(chunk)
        print(f"Download completed: {output_path}")
        return True
    else:
        print(f"Failed to download file. Status code: {response.status_code}")
        return False

def sanitize_filename(name):
    """Convert a string to a valid filename"""
    # Replace spaces and special characters with underscores
    sanitized = re.sub(r'[^\w\-_\. ]', '_', name)
    # Replace multiple consecutive underscores with a single one
    sanitized = re.sub(r'_+', '_', sanitized)
    # Trim to avoid excessively long filenames
    if len(sanitized) > 50:
        sanitized = sanitized[:50]
    return sanitized

def create_folder_structure(base_path, usn_time):
    """Create folder structure for a specific usn_time"""
    # Sanitize usn_time for folder name
    folder_name = sanitize_filename(usn_time)
    item_folder = os.path.join(base_path, folder_name)
    
    # Path mappings to return
    paths = {
        'item_folder': item_folder,
        'parent_folder_name': folder_name  # Store the parent folder name for naming Excel files
    }
    
    # Create all required subfolders
    for subfolder in CONFIG["SUBFOLDERS"]:
        folder_path = os.path.join(item_folder, subfolder)
        os.makedirs(folder_path, exist_ok=True)
        
        # Store the path with normalized slashes
        key_name = subfolder.replace('/', '_') + '_path'
        paths[key_name] = folder_path.replace('\\', '/')
    
    return paths

def process_json_file(api_key, json_file_path):
    """Process a single JSON file to download videos and extract frames"""
    print(f"\n{'='*60}")
    print(f"Processing JSON file: {json_file_path}")
    print(f"{'='*60}")
    
    # Extract JSON filename without extension for creating the base folder
    json_filename = os.path.basename(json_file_path)
    json_name_no_ext = os.path.splitext(json_filename)[0]
    
    # Create base output folder for this JSON file
    json_output_folder = os.path.join(CONFIG["OUTPUT_BASE_FOLDER"], sanitize_filename(json_name_no_ext))
    os.makedirs(json_output_folder, exist_ok=True)
    
    # Initialize Apify client
    client = ApifyClient(api_key)
    
    # Load JSON data
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            loaded_list = json.load(f)
    except Exception as e:
        print(f"Error loading JSON file: {e}")
        return
    
    print(f"Loaded {len(loaded_list)} items from JSON file")
    
    # Process each item in the JSON file
    for idx, item in enumerate(loaded_list):
        # Get usn_time as the identifier for folder structure
        usn_time = item.get('usn_time', '')
        
        if not usn_time:
            print(f"Warning: No usn_time found for item {idx+1}. Using index as identifier.")
            usn_time = f"item_{idx+1}"
        
        print(f"\nProcessing item {idx+1}/{len(loaded_list)} with usn_time: {usn_time}")
        
        # Create folder structure based on usn_time
        paths = create_folder_structure(json_output_folder, usn_time)
        
        # Add paths to JSON item
        item.update({
            'quanan_folder_path': paths['item_folder'].replace('\\', '/'),
            'vid_path': paths['vid_path'],
            'img_path': paths['img_path'],
            'final_imgs_path': paths['final_imgs_path'],
            'cover_img_path': paths['cover_img_path'],
            'comments_path': paths['comments_path'],
            'filter_comments_path': paths['comments_filter_cmt_path'],
            'user_cover_img': paths['comments_user_cover_img_path'],
            # Store parent folder name for Excel files
            'parent_folder_name': paths['parent_folder_name']
        })
        
        # Get TikTok URL from postPage field
        clip_url = item.get('postPage', '')
        
        if not clip_url:
            print(f"Warning: No URL found for {usn_time}. Skipping...")
            continue
        
        print(f"TikTok URL: {clip_url}")
        
        # Download the video using Apify
        run_input = {
            "postURLs": [clip_url],
            "shouldDownloadVideos": True,
            "shouldDownloadCovers": True,
            "shouldDownloadSubtitles": False,
            "shouldDownloadSlideshowImages": False,
        }
        
        try:
            print(f"Calling Apify API to download video...")
            run = client.actor("S5h7zRLfKFEr8pdj7").call(run_input=run_input)
            print(f"Apify run completed, dataset ID: {run['defaultDatasetId']}")
            media_urls = None
            
            # Get download URLs from the dataset
            dataset_items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
            print(f"Dataset contains {len(dataset_items)} items")
            
            if dataset_items:
                first_item = dataset_items[0]
                print(f"Dataset item structure: {list(first_item.keys())}")
                
                # Look for media URLs in different possible field names
                possible_fields = ['mediaUrls', 'videoUrl', 'videoUrls', 'urls', 'video']
                
                for field in possible_fields:
                    if field in first_item:
                        media_urls = first_item[field]
                        print(f"Found media URLs in field '{field}': {type(media_urls)}")
                        break
                        
                if not media_urls and 'video' in first_item:
                    media_urls = [first_item['video']]
                
                # Try to get cover image URL
                if 'cover' in first_item:
                    cover_url = first_item['cover']
                    cover_filename = f"{sanitize_filename(usn_time)}_cover.jpg"
                    cover_path = os.path.join(paths['cover_img_path'], cover_filename)
                    
                    try:
                        success = download_mp4(cover_url, cover_path)
                        if success:
                            item['cover_img'] = cover_path.replace('\\', '/')
                    except Exception as e:
                        print(f"Error downloading cover image: {e}")
            
            if not media_urls:
                print(f"No media URLs found for {usn_time}. Skipping...")
                continue
            
            # Update the item with download URL
            loaded_list[idx]['downloadUrl'] = media_urls
            
            # Handle different formats of mediaUrls
            video_url = ''
            if isinstance(media_urls, dict):
                video_url = media_urls.get('video', '')
            elif isinstance(media_urls, list):
                for url in media_urls:
                    if isinstance(url, str) and (url.endswith('.mp4') or 'video' in url):
                        video_url = url
                        break
                if not video_url and media_urls:
                    video_url = media_urls[0]
            elif isinstance(media_urls, str):
                video_url = media_urls
            
            if video_url:
                video_filename = f"{sanitize_filename(usn_time)}.mp4"
                video_path = os.path.join(paths['vid_path'], video_filename)
                
                success = download_mp4(video_url, video_path)
                item['video_file'] = video_path.replace('\\', '/')
                
                if success:
                    # Extract frames and save paths
                    frame_paths = extract_frames(
                        video_path, 
                        paths['img_path'], 
                        CONFIG["FRAME_INTERVAL"]
                    )
                    item['frames'] = [path.replace('\\', '/') for path in frame_paths]
                    print(f"Extracted {len(frame_paths)} frames to {paths['img_path']}")
                
                # Create Excel file for comments with parent folder name
                # Name the file after the parent folder (usn_time folder)
                comments_excel_filename = f"{paths['parent_folder_name']}.xlsx"
                comments_excel_path = os.path.join(paths['comments_path'], comments_excel_filename)
                
                # Create an empty Excel file as a placeholder
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Comments Placeholder"
                    ws['A1'] = "This is a placeholder for TikTok comments."
                    ws['A2'] = f"TikTok URL: {clip_url}"
                    ws['A3'] = "Comments will be extracted in a separate process."
                    wb.save(comments_excel_path)
                    print(f"Created comments placeholder Excel file: {comments_excel_path}")
                    
                    # Add comments Excel file path to JSON
                    item['comments_excel'] = comments_excel_path.replace('\\', '/')
                except Exception as e:
                    print(f"Error creating comments Excel placeholder: {e}")
                
        except Exception as e:
            print(f"Error processing {usn_time}: {e}")
    
    # Save updated JSON with download URLs and media paths
    output_json = os.path.join(
        json_output_folder, 
        f"{os.path.basename(json_file_path).replace('.json', '')}_processed.json"
    )
    
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(loaded_list, f, ensure_ascii=False, indent=4)
    
    print(f"Updated JSON saved to {output_json}")
    return output_json

def batch_process_json_files(input_folder):
    """Process all JSON files in the input folder"""
    # Ensure output base folder exists
    os.makedirs(CONFIG["OUTPUT_BASE_FOLDER"], exist_ok=True)
    
    # Find all JSON files in the input folder and its subfolders
    json_pattern = os.path.join(input_folder, "**", "*.json")
    json_files = glob.glob(json_pattern, recursive=True)
    
    if not json_files:
        print(f"No JSON files found in {input_folder}")
        return
    
    print(f"Found {len(json_files)} JSON files to process")
    
    # Process each JSON file
    processed_files = []
    for i, json_file in enumerate(json_files):
        print(f"\nProcessing file {i+1}/{len(json_files)}: {json_file}")
        output_file = process_json_file(CONFIG["API_KEY"], json_file)
        if output_file:
            processed_files.append(output_file)
    
    print(f"\nBatch processing complete. Processed {len(processed_files)} JSON files.")
    for file in processed_files:
        print(f"- {file}")

def main():
    print("Starting batch processing of JSON files")
    print(f"Input folder: {CONFIG['INPUT_FOLDER']}")
    print(f"Output base folder: {CONFIG['OUTPUT_BASE_FOLDER']}")
    
    batch_process_json_files(CONFIG["INPUT_FOLDER"])

if __name__ == "__main__":
    main()