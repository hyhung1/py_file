import json
import os
import pandas as pd
import traceback
from apify_client import ApifyClient
from typing import Dict, List, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ============ Configuration Variables (All in one place) ============
CONFIG = {
    # API credentials
    "API_KEY": "apify_api_aiwzjla0jREK3paH0j7cmhrZlgjPWB0mXFRe",
    
    # TikTok search parameters
    "MAX_ITEMS": 10,
    "SEARCH_LOCATION": "VN",
    "DATE_RANGE": "DEFAULT",
    
    # File paths
    "INPUT_JSON_PATH": "./QUANAN_alpha/q10/quán_ngon_quận_10.json",
    "INPUT_EXCEL_PATH": "./QUANAN_alpha/q10/quán_ngon_quận_10_v2.xlsx",
    "OUTPUT_DIR": "cac_quanan_q10/json_xlsx",
    
    # Excel configuration
    "EXCEL_HEADERS": ["usn_time", "postPage", "title"],
    "HEADER_COLOR": "DDEBF7"
}

def load_existing_data(json_file_path: str) -> List[Dict]:
    """Load existing restaurant data from JSON file"""
    with open(json_file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def search_tiktok_videos(search_term: str, max_items: int) -> List[Dict[str, Any]]:
    """Search TikTok videos based on a search term using Apify API"""
    # Initialize the ApifyClient
    client = ApifyClient(CONFIG["API_KEY"])
    
    # Prepare the Actor input
    run_input = {
        "maxItems": max_items,
        "keywords": [search_term],
        "dateRange": CONFIG["DATE_RANGE"],
        "location": CONFIG["SEARCH_LOCATION"],
        "customMapFunction": "(object) => { return {...object} }"
    }
    
    print(f"Searching TikTok for: {search_term}")
    print(f"Maximum items: {max_items}")
    
    # Run the Actor and wait for it to finish
    run = client.actor("5K30i8aFccKNF5ICs").call(run_input=run_input)
    
    results = []
    
    # Fetch and process Actor results
    print("Processing search results...")
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        # Extract only the requested fields
        extracted_data = {
            "title": item.get("title"),
            "views": item.get("views"),
            "likes": item.get("likes"),
            "comments": item.get("comments"),
            "shares": item.get("shares"),
            "bookmarks": item.get("bookmarks"),
            "hashtags": item.get("hashtags"),
            "uploadedAt": item.get("uploadedAt"),
            "uploadedAtFormatted": item.get("uploadedAtFormatted"),
            "channel": {
                "name": item.get("channel", {}).get("name"),
                "username": item.get("channel", {}).get("username")
            },
            "postPage": item.get("postPage")
        }
        
        # Format date and create identifier (usn_time)
        upload_time = item.get("uploadedAt")
        if upload_time:
            dt = datetime.fromtimestamp(upload_time)
            formatted_date = dt.strftime('%Y_%d_%m')
            channel_name = extracted_data["channel"]["username"]
            extracted_data["usn_time"] = f"{channel_name}_{formatted_date}"
        
        results.append(extracted_data)
    
    return results

def create_excel_file(data: List[Dict], filename: str):
    """Create an Excel file with TikTok video data"""
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TikTok Data"
    
    # Use headers from config
    headers = CONFIG["EXCEL_HEADERS"]
    
    # Style for header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color=CONFIG["HEADER_COLOR"], 
                             end_color=CONFIG["HEADER_COLOR"], 
                             fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, video in enumerate(data, 2):
        # Fill in only the requested fields
        ws.cell(row=row_num, column=1, value=video.get("usn_time", ""))
        ws.cell(row=row_num, column=2, value=video.get("postPage", ""))
        ws.cell(row=row_num, column=3, value=video.get("title", ""))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    print(f"Created Excel file: {filename}")

def extract_excel_data(xlsx_file):
    """Extract restaurant data from Excel file with usn_time as key"""
    try:
        print(f"\nExtracting data from Excel file: {xlsx_file}")
        if not os.path.exists(xlsx_file):
            print(f"Error: Excel file '{xlsx_file}' does not exist.")
            return {}
            
        df = pd.read_excel(xlsx_file)
        print(f"DataFrame shape: {df.shape}")
        
        restaurant_data = {}
        
        for idx, row in df.iterrows():
            usn_time = row.get('usn_time')
            
            if not usn_time:
                continue
            
            # Get eat_name and remove trailing spaces while preserving internal spaces
            eat_name = row.get('eat_name', '')
            if isinstance(eat_name, str):
                eat_name = eat_name.rstrip()
                
            restaurant_data[usn_time] = {
                'eat_name': eat_name,
                'eat_addr': row.get('eat_addr', ''),
                'open_time': row.get('open_time', ''),
                'menu': row.get('menu', '')
            }
            
        print(f"Total restaurant entries processed: {len(restaurant_data)}")
        return restaurant_data
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        print(traceback.format_exc())
        return {}

def update_json_data(json_file, restaurant_data):
    """Update JSON file with restaurant details based on matching usn_time"""
    try:
        print(f"\nUpdating JSON data from: {json_file}")
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        updated_count = 0
        removed_count = 0
        filtered_data = []
        
        for entry in data:
            usn_time = entry.get('usn_time', '')
            
            if usn_time and usn_time in restaurant_data:
                entry.update(restaurant_data[usn_time])
                filtered_data.append(entry)
                updated_count += 1
            else:
                removed_count += 1
        
        # Create the updated filename with _upd suffix
        file_name, file_ext = os.path.splitext(json_file)
        new_file = f"{file_name}_upd{file_ext}"
        
        with open(new_file, 'w', encoding='utf-8') as f:
            json.dump(filtered_data, f, ensure_ascii=False, indent=4)
        
        print(f"Updated and kept {updated_count} entries")
        print(f"Removed {removed_count} entries")
        print(f"Saved to {new_file}")
        return new_file
        
    except Exception as e:
        print(f"Error updating JSON file: {e}")
        print(traceback.format_exc())
        return None

# Task 1: Update JSON with Excel data
def task1_update_json():
    """Update existing JSON data with restaurant details from Excel"""
    print("\n===== TASK 1: UPDATE JSON WITH EXCEL DATA =====")
    
    input_json_path = CONFIG["INPUT_JSON_PATH"]
    excel_path = CONFIG["INPUT_EXCEL_PATH"]
    
    if not os.path.exists(input_json_path):
        print(f"Error: JSON file '{input_json_path}' does not exist.")
        return None
        
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' does not exist.")
        return None
    
    restaurant_data = extract_excel_data(excel_path)
    
    if not restaurant_data:
        print("No restaurant data found in Excel file.")
        return None
    
    updated_json_path = update_json_data(input_json_path, restaurant_data)
    print(f"Task 1 completed. Updated JSON file: {updated_json_path}")
    return updated_json_path

# Task 2: Search TikTok for restaurants from updated JSON
def task2_search_tiktok(updated_json_path):
    """Search TikTok for videos related to restaurants from updated JSON"""
    print("\n===== TASK 2: SEARCH TIKTOK FOR RESTAURANTS =====")
    
    output_dir = CONFIG["OUTPUT_DIR"]
    max_items = CONFIG["MAX_ITEMS"]
    
    if not updated_json_path or not os.path.exists(updated_json_path):
        print(f"Error: Updated JSON file not found.")
        return False
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Load restaurant data from updated JSON
    print(f"Loading data from: {updated_json_path}")
    restaurants = load_existing_data(updated_json_path)
    
    if not restaurants:
        print("No restaurant data found in updated JSON file.")
        return False
    
    print(f"Found {len(restaurants)} restaurants in the updated JSON.")
    
    # Process each restaurant
    for restaurant in restaurants:
        eat_name = restaurant.get("eat_name")
        if eat_name and eat_name != "NaN" and not isinstance(eat_name, float):
            print(f"\nProcessing restaurant: {eat_name}")
            
            # Search TikTok for videos related to this restaurant
            videos = search_tiktok_videos(eat_name, max_items)
            
            # Create safe filename
            safe_name = eat_name.replace("/", "_").replace("\\", "_").replace(":", "_")\
                              .replace("*", "_").replace("?", "_").replace("\"", "_")\
                              .replace("<", "_").replace(">", "_").replace("|", "_")\
                              .replace(" ", "_")
            
            # Save all restaurant videos to a single JSON file
            json_filename = f"{output_dir}/{safe_name}.json"
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(videos, f, ensure_ascii=False, indent=4)
            print(f"Created JSON file: {json_filename}")
            
            # Create Excel file with the same data
            excel_filename = f"{output_dir}/{safe_name}.xlsx"
            create_excel_file(videos, excel_filename)
            
            print(f"Processed {len(videos)} videos for {eat_name}")
    
    print("\nTask 2 completed. All restaurants processed successfully.")
    return True

def main():
    # Print current working directory
    print(f"Current working directory: {os.getcwd()}")
    
    # Task 1: Update JSON with Excel data
    updated_json_path = task1_update_json()
    
    if updated_json_path:
        # Task 2: Search TikTok for restaurants
        task2_search_tiktok(updated_json_path)
    else:
        print("Cannot proceed to Task 2 because Task 1 failed.")

if __name__ == "__main__":
    main()