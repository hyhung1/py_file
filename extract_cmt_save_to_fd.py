from apify_client import ApifyClient
from datetime import datetime
import openpyxl
import json
import os
import sys

def extract_tiktok_comments(api_key, url, max_items=80, output_file=None):
    """
    Extract comments from a TikTok video and save them to an Excel file.
    
    Args:
        api_key (str): Apify API key
        url (str): URL of the TikTok video
        max_items (int): Maximum number of comments to extract (default: 80)
        output_file (str): Name of the output Excel file
    
    Returns:
        int, list: Number of comments extracted and the extracted data
    """
    # Initialize the ApifyClient with the provided API token
    client = ApifyClient(api_key)

    # Prepare the Actor input
    run_input = {
        "startUrls": [url],
        "includeReplies": True,
        "maxItems": max_items,
        "customMapFunction": "(object) => { return {...object} }",
    }

    # Run the Actor and wait for it to finish
    print(f"Extracting comments from {url}...")
    run = client.actor("XomSRf7d0qf3mVj1y").call(run_input=run_input)

    # Fetch and extract specific fields from each item
    extracted_data = []
    print("Processing comments...")
    for item in client.dataset(run["defaultDatasetId"]).iterate_items():
        # Parse the date string and reformat it
        created_at = item.get("createdAt", "")
        formatted_date = ""
        if created_at:
            # Parse ISO format date
            date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            # Format as DD-MM-YYYY
            formatted_date = date_obj.strftime('%d-%m-%Y')
        
        # Extract only the fields we need
        filtered_item = {
            "text": item.get("text"),
            "createdAt": formatted_date,  # Use the reformatted date
            "likeCount": item.get("likeCount"),
            "replyCount": item.get("replyCount"),
            "isAuthorLiked": item.get("isAuthorLiked")
        }
        
        # Extract user information if it exists
        if "user" in item and item["user"]:
            filtered_item["username"] = item["user"].get("username")
            filtered_item["displayName"] = item["user"].get("displayName")
            filtered_item["bio"] = item["user"].get("bio")
            filtered_item["avatarUrl"] = item["user"].get("avatarUrl")
        
        extracted_data.append(filtered_item)

    # Only create Excel file if there's data and output_file is specified
    if extracted_data and output_file:
        # Create a new Excel workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "TikTok Comments"

        # Define headers
        headers = ["Text", "Created At", "Like Count", "Reply Count", "Is Author Liked", 
                  "Username", "Display Name", "Bio", "Avatar URL"]

        # Write headers to the first row
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col).value = header

        # Write data to the sheet
        for row, data in enumerate(extracted_data, 2):  # Start from row 2 (after headers)
            sheet.cell(row=row, column=1).value = data.get("text", "")
            sheet.cell(row=row, column=2).value = data.get("createdAt", "")
            sheet.cell(row=row, column=3).value = data.get("likeCount", 0)
            sheet.cell(row=row, column=4).value = data.get("replyCount", 0)
            sheet.cell(row=row, column=5).value = "Yes" if data.get("isAuthorLiked") else "No"
            sheet.cell(row=row, column=6).value = data.get("username", "")
            sheet.cell(row=row, column=7).value = data.get("displayName", "")
            sheet.cell(row=row, column=8).value = data.get("bio", "")
            sheet.cell(row=row, column=9).value = data.get("avatarUrl", "")

        # Auto-adjust column widths to fit content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50

        # Print the absolute path of the output file
        abs_path = os.path.abspath(output_file)
        print(f"EXCEL FILE PATH: {abs_path}")
        
        # Ensure directory exists
        dirname = os.path.dirname(output_file)
        if dirname:  # Only try to create directory if there is a directory component
            print(f"CREATING DIRECTORY: {os.path.abspath(dirname)}")
            os.makedirs(dirname, exist_ok=True)
        
        # Save the workbook
        try:
            workbook.save(output_file)
            print(f"Data saved to {output_file}")  # Keep this to match your original output
        except Exception as e:
            print(f"ERROR SAVING EXCEL FILE: {str(e)}")
            raise
    
    return len(extracted_data), extracted_data


def process_json_file(json_file, api_key, max_comments=80):
    """
    Process all restaurants in the JSON file:
    1. Check if comments folder exists for each restaurant
    2. Extract TikTok comments for each restaurant's video
    3. Save the comments to Excel files in the comments folders
    
    Args:
        json_file (str): Path to the JSON file containing restaurants data
        api_key (str): Apify API key
        max_comments (int): Maximum number of comments to extract per video
    """
    try:
        # Parse the JSON data
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        print(f"Found {len(data)} restaurants in the JSON file")
        
        # Loop through each item in the JSON array
        for index, item in enumerate(data):
            print(f"\nProcessing restaurant {index+1}/{len(data)}")
            
            try:
                # Get restaurant details
                restaurant_name = item.get('eat_name', 'Unknown')
                print(f"Restaurant: {restaurant_name}")
                
                # Extract media_path
                media_path = item.get('media_path')
                if not media_path:
                    print(f"  ERROR: Media path not found for {restaurant_name}")
                    continue
                
                print(f"  Media path: {media_path}")
                
                # Concatenate media_path with 'comments'
                comments_path = os.path.join(media_path, 'comments')
                print(f"  Comments path: {comments_path}")
                
                # Check if the comments directory exists, create if it doesn't
                if not os.path.exists(comments_path):
                    try:
                        os.makedirs(comments_path, exist_ok=True)
                        print(f"  Created comments directory: {comments_path}")
                    except Exception as e:
                        print(f"  ERROR: Could not create comments directory: {str(e)}")
                        continue
                else:
                    print(f"  Comments directory already exists")
                
                # Extract the TikTok video URL
                post_page = item.get('postPage')
                if not post_page:
                    print(f"  ERROR: TikTok URL (postPage) not found for {restaurant_name}")
                    continue
                
                print(f"  TikTok URL: {post_page}")
                
                # Generate a filename for the Excel file
                safe_name = restaurant_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                channel_username = item.get('channel', {}).get('username', 'unknown')
                
                # Ensure we use the correct filename and not a default one
                excel_filename = f"{safe_name}_cmt.xlsx"
                excel_path = os.path.join(comments_path, excel_filename)
                
                print(f"  Will save comments to: {excel_path}")
                
                # Extract comments and save to Excel - pass FULL PATH to output_file
                comments_count, _ = extract_tiktok_comments(
                    api_key=api_key,
                    url=post_page,
                    max_items=max_comments,
                    output_file=excel_path  # This must be the full path
                )
                
                print(f"  SUCCESS: Extracted {comments_count} comments for {restaurant_name}")
            
            except Exception as e:
                print(f"  ERROR processing restaurant {restaurant_name}: {str(e)}")
        
        print("\nProcessing complete!")
    
    except Exception as e:
        print(f"ERROR: Failed to process JSON file: {str(e)}")


# Main execution
if __name__ == "__main__":
    # Replace these with your actual values
    API_KEY = "apify_api_aiwzjla0jREK3paH0j7cmhrZlgjPWB0mXFRe"  # Your Apify API key
    JSON_FILE = "./QUANAN/q10/quán_ngon_quận_10_upd_addurl.json"  # Your JSON file path
    MAX_COMMENTS = 20  # Maximum number of comments to extract per video
    
    # Print current working directory
    print(f"Current working directory: {os.getcwd()}")
    
    # Check if JSON file exists
    if not os.path.exists(JSON_FILE):
        print(f"ERROR: JSON file '{JSON_FILE}' not found")
        sys.exit(1)
    else:
        print(f"JSON file found: {os.path.abspath(JSON_FILE)}")
    
    # Process all restaurants in the JSON file
    process_json_file(
        json_file=JSON_FILE,
        api_key=API_KEY,
        max_comments=MAX_COMMENTS
    )